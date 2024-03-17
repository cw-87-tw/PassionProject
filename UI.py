import tkinter as tk
from tkinter import ttk
from getData import *
try:
    import openpyxl
except:
    import os
    os.system("pip install openpyxl")
    import openpyxl
root = tk.Tk()

print("start TK")

def init():
    root.title("Passion 財務報表查詢")
    root.geometry("800x800")

    for i in range(3):
        root.rowconfigure(i, weight=1)
    root.rowconfigure(3, weight = 3)
    for i in range(2):
        root.columnconfigure(i, weight=1)

    global entry_target
    global entry_year
    global entry_season
    # global entry_user
    label_target = tk.Label(root, text="輸入查詢編號:", font=("Microsoft JhengHei UI", 16))  # 設置字型大小為12
    entry_target = tk.Entry(root, font=("Microsoft JhengHei UI", 16))
    label_year = tk.Label(root, text="查詢年份:", font=("Microsoft JhengHei UI", 16))  # 設置字型大小為12
    entry_year = tk.Entry(root, font=("Microsoft JhengHei UI", 16))
    label_season = tk.Label(root, text="查詢季度:", font=("Microsoft JhengHei UI", 16))  # 設置字型大小為12
    entry_season = tk.Entry(root, font=("Microsoft JhengHei UI", 16))
    # label_user = tk.Label(root, text="使用者:", font=("Microsoft JhengHei UI", 16))  # 設置字型大小為12
    # entry_user= tk.Entry(root, font=("Microsoft JhengHei UI", 16))
    global button_search
    button_search = tk.Button(root, text="查詢", font=("Microsoft JhengHei UI", 16), command=search)
    global tree
    tree = ttk.Treeview(root, columns=("Attribute", "Value"), show="headings")
    tree.heading("Attribute", text="Attribute")
    tree.heading("Value", text="Value")
    global previousResult
    previousResult = list()

    # label_target.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
    # entry_target.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
    label_year.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
    entry_year.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
    label_season.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
    entry_season.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
    # label_user.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")
    # entry_user.grid(row=2, column=1, padx=10, pady=10, sticky="nsew")
    button_search.grid(row=2, columnspan=2, padx=10, pady=10, sticky="nsew")
    tree.grid(row=3, columnspan=2, padx=10, pady=10, sticky="nsew")
    
    # 設置Treeview文字大小
    style = ttk.Style()
    style.configure("Treeview", font=("Microsoft JhengHei UI", 14))  # 設置字型大小為14

def changeButton(text, state):
    button_search.config(text=text, state = state)

def show_result(result):
    global previousResult
    print(previousResult)
    for i in previousResult: tree.delete(i)
    previousResult.clear()
    for key, value in result.items():
        previousResult.append(tree.insert("", "end", values=(key, value)))
    print(previousResult)
    root.update()
    # previousResult = result


def read_stock_numbers_from_excel(filename):
    stock_numbers = []
    readwb = openpyxl.load_workbook(filename)
    readws = readwb.active
    for row in readws.iter_rows(values_only=True):
        stock_numbers.append(row[2])
    # print("stock numbers:", *stock_numbers)
    print("got stock numbers", len(stock_numbers))
    return stock_numbers

wb = openpyxl.Workbook()
ws = wb.active
l = []
def save_to_excel(results : dict, name):
    if len(l) == 0: ws.append(["公司編號"] + list(results.keys()))  # 添加標題行
    l.append([name] + list(results.values()))
    name = len(l)
    ws.append([name] + list(results.values()))
    wb.save(f"{entry_year.get()}年第{entry_season.get()}季.xlsx")

def search():
    changeButton(text="搜尋中...請稍候", state = "disabled")
    root.update()
    l.clear()
    global wb
    wb = openpyxl.Workbook()
    global ws
    ws = wb.active
    stock_numbers = read_stock_numbers_from_excel("stocks.xlsx")
    cnt = 1
    login("h1110539@stu.wghs.tp.edu.tw")
    for stock in stock_numbers:
        setInformation(entry_year.get(), entry_season.get(), stock)
        while 1:
            try:
                results = dict()
                results.update(getPrice())
                results.update(getIncome())
                results.update(getCash())
                results.update(getDebt())
                results.update(getRoeRoa())
                results.update(getAssets())
                results.update(getEps())
                save_to_excel(results, stock)
                break
            except: pass
        show_result({"目前進度" : cnt})
        print("目前進度:", cnt)
        cnt += 1
    changeButton(text="查詢", state = "normal")
    show_result(result={"狀態" : "成功"})
    