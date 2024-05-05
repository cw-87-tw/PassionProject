import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import openpyxl
import tkinter.messagebox as mb
import os

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("股票購買與賣出分析工具")

        self.buy_sheet = ""
        self.output_name = "購買分析結果"
        self.money = 5000000

        self.label1 = ttk.Label(root, text="欲模擬購買來源:")
        self.label1.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.buy_sheet_entry = ttk.Entry(root)
        self.buy_sheet_entry.grid(row=0, column=1, padx=5, pady=5)
        
        self.browse_button = ttk.Button(root, text="瀏覽...", command=self.browse_buy_sheet)
        self.browse_button.grid(row=0, column=2, padx=5, pady=5)

        self.label2 = ttk.Label(root, text="輸出檔案名稱:")
        self.label2.grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.output_name_entry = ttk.Entry(root)
        self.output_name_entry.grid(row=1, column=1, padx=5, pady=5)
        self.output_name_entry.insert(0, self.output_name)

        self.label3 = ttk.Label(root, text="初始資金:")
        self.label3.grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.money_entry = ttk.Entry(root)
        self.money_entry.grid(row=2, column=1, padx=5, pady=5)
        self.money_entry.insert(0, self.money)

        self.run_button = ttk.Button(root, text="執行", command=self.run)
        self.run_button.grid(row=3, column=1, padx=5, pady=5)

    def browse_buy_sheet(self):
        self.buy_sheet = filedialog.askopenfilename()
        self.buy_sheet_entry.delete(0, tk.END)
        self.buy_sheet_entry.insert(0, self.buy_sheet)

    def run(self):
        self.buy_sheet = self.buy_sheet_entry.get()
        self.output_name = self.output_name_entry.get()
        self.money = float(self.money_entry.get())
        run(self.buy_sheet, self.output_name, self.money)

def getBuy(buySheet) -> dict:
    wb = openpyxl.load_workbook(buySheet)
    ws = wb.active

    allTargets = dict()
    for row in ws.iter_rows(values_only=True, min_row=2):
        allTargets[row[0]] = list()
        for cell in row[1:]:
            if cell is not None:
                allTargets[row[0]].append(cell)
    return allTargets

def getPrices(year: int) -> dict:
    readwb = openpyxl.load_workbook(f"./results/{year}年.xlsx")
    readws = readwb.active
    idx = 1
    prices = dict()
    for col in readws.iter_cols(values_only=True, min_col=2, max_col=2):
        for i in col[1:]:
            prices[idx] = i if type(i) != str else 1e20
            idx += 1
    return prices

def buyAndSell(allTargets: dict, outputName: str, money: float) -> None:
    hold = [] # (number, 股數)
    wb = openpyxl.Workbook()
    ws = wb.active

    for year, targets in allTargets.items():
        prices = getPrices(int(year)) # get this year's prices

        # sell old
        for no, shares in hold:
            if prices[no] == 1e20: raise Exception(f"{year}年無法賣出此股票(可能因為下市等因素)\n錯誤編號為{no}")
            money += shares * prices[no]
        hold.clear()

        ws.append([f"{year}年"])
        ws.append(["原有資金", money])
        ws.append([])
        ws.append(["購買編號", "股價", "股數", "總價"])

        # buy new
        perMoney = money / len(targets)
        for stock in targets:
            if prices[stock] == 1e20: raise Exception(f"{year}年無法買入此股票(可能因為下市等因素)\n錯誤編號為{stock}")
            shares = perMoney // prices[stock]
            money -= prices[stock] * shares
            ws.append([stock, prices[stock], shares, prices[stock] * shares])
            hold.append((stock, shares))

        ws.append([])
        ws.append(["剩餘資金", money])
        ws.append([])
        ws.append([])
        ws.append([])

    lastYear = list(allTargets.keys())[-1] + 1
    prices = getPrices(lastYear) # get this year's prices
    # sell old
    for no, shares in hold:
        if prices[no] == 1e20: raise Exception(f"{lastYear}年無法賣出此股票(可能因為下市等因素)\n錯誤編號為{no}")
        money += shares * prices[no]
    hold.clear()

    ws.append(["最終資金", money])
        
    saveDir = "analysisResults"
    if saveDir not in os.listdir(): os.mkdir(saveDir)
    wb.save(f"./{saveDir}/{outputName}.xlsx")
    mb.showinfo("成功", f"檔案成功儲存: {os.getcwd()}/{saveDir}/{outputName}.xlsx")

def run(buySheet, outputName, money):
    try:
        allTargets = getBuy(buySheet)
        buyAndSell(allTargets, outputName, money)
    except Exception as e:
        mb.showerror("失敗", f"失敗，錯誤訊息:\n{e}")
        print(e)

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
