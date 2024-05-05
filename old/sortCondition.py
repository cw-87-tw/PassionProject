import openpyxl

beginYear = 2013
# beginYear = int(input("beginYear:"))
endYear = 2022
# endYear = int(input("endYear:"))
factor = "營業利益率"
highest = True
numbers = 5

def getData() -> dict:
    results = dict()
    for year in range(beginYear, endYear + 1):
        results[year] = dict()
        readwb = openpyxl.load_workbook(f"./results/{year}年.xlsx")
        readws = readwb.active

        for col in readws.iter_cols(values_only = True, min_col = 2):
            factor = col[0]
            col = col[1:]
            for idx, i in enumerate(col, 1): # 到時候可能不能用 enumerate 因為要用真的股票編號
                if results[year].get(idx) == None: results[year][idx] = dict()
                try: results[year][idx][factor] = float(i)
                except: results[year][idx][factor] = "Error"
                # print(idx, prices[idx])
    return results

def customSort(results: dict, year: int, factor: str, highest: bool = False, numbers: int = 5,) -> list: # number: 要挑出最好的幾個?
    dataList = [(stock, data) for stock, data in results[year].items()]
    def customKey(item: tuple): # 用來排序的判準
        if type(item[1][factor]) != float or item[1]["股價"] == "Error":
            # print(item)
            if highest: return -1e20
            else: return 1e20
        return item[1][factor]
    dataList.sort(key = customKey, reverse = highest) # reverse: False 小到大, True 大到小
    buyList = list()
    for stock, data in dataList[:numbers]:
        buyList.append(stock)
    return buyList

wb = openpyxl.Workbook()
ws = wb.active
def addData(buyList: list, year: int) -> None:
    if year == beginYear: ws.append(["年份", "以下為購買的對象編號"])
    ws.append([year] + buyList)


results = getData()
# print(results)
for year in range(beginYear, endYear + 1):
    buyList = customSort(results, year, factor = factor, highest = highest, numbers = numbers)
    addData(buyList, year)


wb.save(f"購買標的({factor}).xlsx")