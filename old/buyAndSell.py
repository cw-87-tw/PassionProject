import openpyxl
from datetime import datetime

# should be able to be modified my users
buySheet = input("excel 檔案名稱:")
outputName = "購買分析結果"
money = 5000000
# end here

def getBuy() -> dict:
    wb = openpyxl.load_workbook(buySheet)
    ws = wb.active

    allTargets = dict()
    for row in ws.iter_rows(values_only=True, min_row = 2):
        allTargets[row[0]] = list()
        for cell in row[1:]:
            if cell != None: allTargets[row[0]].append(cell)
    return allTargets


def getPrices(year: int) -> dict:
    readwb = openpyxl.load_workbook(f"./results/{year}年.xlsx")
    readws = readwb.active
    idx = 1
    prices = dict()
    for col in readws.iter_cols(values_only = True, min_col = 2, max_col = 2):
        for i in col[1:]:
            prices[idx] = i if type(i) != str else 1e20
            # print(idx, prices[idx])
            idx += 1
    return prices

def buyAndSell(allTargets: dict) -> None:
    hold = [] # (number, 股數)
    wb = openpyxl.Workbook()
    ws = wb.active

    for year, targets in allTargets.items():
        prices = getPrices(year) # get this year's prices

        # sell old
        for no, shares in hold:
            if prices[no] == 1e20: raise Exception(f"{year}年無法賣出此股票(可能因為下市等因素)\n錯誤編號為{no}")
            money += shares * prices[no]
        hold.clear()

        ws.append([f"{year}年"])
        ws.append(["原有資金", money])
        ws.append([])
        ws.append(["購買編號", "股價", "股數", "總價"])

        # buy new
        perMoney = money / len(targets)
        for stock in targets:
            if prices[stock] == 1e20: raise Exception(f"{year}年無法買入此股票(可能因為下市等因素)\n錯誤編號為{stock}")
            shares = perMoney // prices[stock]
            money -= prices[stock] * shares
            ws.append([stock, prices[stock], shares, prices[stock] * shares])
            print(f"{year}購買了{stock}，買了{shares}股")
            hold.append((stock, shares))

        ws.append([])
        ws.append(["剩餘資金", money])
        ws.append([])
        ws.append([])
        ws.append([])


    lastYear = list(allTargets.keys())[-1] + 1
    prices = getPrices(lastYear) # get this year's prices
    # sell old
    for no, shares in hold:
        if prices[no] == 1e20: raise Exception(f"{lastYear}年無法賣出此股票(可能因為下市等因素)\n錯誤編號為{no}")
        money += shares * prices[no]
    hold.clear()

    ws.append(["最終資金", money])
        
    wb.save(f"./{outputName}.xlsx")

def run():
    try:
        allTargets = getBuy()
        buyAndSell(allTargets)
        print(f"成功，檔案儲存在{outputName}")
    except Exception as e:
        print(f"失敗，錯誤訊息:\n{e}")

run()