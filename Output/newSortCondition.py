import openpyxl

# the following can be modified by user
beginYear = 2013
endYear = 2022
factors = [("營業利率", 1)] # 每個 factor: (name, 要算幾份)
numbers = 5
# end

def getData() -> dict:
    """
    the function can get data from the results crawled on the Net
    the workbook must be located at the desinated directory(./results/)
    it would get all the data in the following format
    results = { 
        beginYear: {
            1: {
                factor1: 123,
                factor2: "Error"...
            },
            2: {
            
            }...
        }...
    }
    """
    results = dict()
    for year in range(beginYear, endYear + 1):
        results[year] = dict()
        readwb = openpyxl.load_workbook(f"./results/{year}年.xlsx")
        readws = readwb.active

        for col in readws.iter_cols(values_only = True, min_col = 2):
            factor = col[0]
            col = col[1:]
            for idx, i in enumerate(col, 1): # 到時候可能不能用 enumerate 因為要用真的股票編號
                if results[year].get(idx) == None: results[year][idx] = dict()
                try: results[year][idx][factor] = float(i)
                except: results[year][idx][factor] = "Error"
                # print(idx, prices[idx])
    return results

def customSort(results: dict, year: int, factors: list, numbers: int = 5) -> list:
    """
    This function can return the best stock numbers to buy
    parameters:
        results: results from getData function
        year: the target year
        factors: a list of factor, each factor is formatted like (factorName, multiple)
    the sorting function will calculate the points for each stock, and return the top {numbers} stocks
    the points would be calculated by finding the index, the from {len(stocks)} to 1, multiply the {multiple} of the factor
    """
    dataList = [(stock, data) for stock, data in results[year].items()]
    
    indexed = dict()
    for factor, multiple in factors:
        tmp = list()
        for stock, data in dataList:
            if type(data[factor]) != str: tmp.append(data[factor])
        indexed[factor] = sorted(tmp)

    def customKey(item: tuple): # 用來排序的判準
        key = 0
        for factor, multiple in factors:
            if type(item[1][factor]) != str and type(item[1]["股價"]) != str: key += indexed[factor].index(item[1][factor]) * multiple
            else: key += (-1e20) * multiple
        return key

    dataList.sort(key = customKey, reverse = True) # 大到小
    buyList = list()
    for stock, data in dataList[:numbers]:
        buyList.append(stock)
    return buyList

class Sheet:
    def __init__(self) -> None:
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def addData(self, buyList: list, year: int) -> None:
        global beginYear
        if year == beginYear: self.ws.append(["年份", "以下為購買的對象編號"])
        self.ws.append([year] + buyList)
    
    def save(self, filename: str):
        self.wb.save(f"{filename}.xlsx")

results = getData()
output = Sheet()
for year in range(beginYear, endYear + 1):
    buyList = customSort(results, year, factors = factors, numbers = numbers) # 每年購買
    output.addData(buyList, year) # add new results to sheet

outputList = list()
for factor, multiple in factors: outputList.append(f"{factor} x {multiple}")
output.save(f"./buyResults/購買標的({','.join(outputList)}).xlsx")