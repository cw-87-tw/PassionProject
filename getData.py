from selenium import webdriver
from twstock import Stock
import openpyxl
from time import sleep
from selenium.webdriver.firefox.options import Options
import os

options = Options()
# options.add_argument("--headless") # 隱藏瀏覽器視窗
driver = webdriver.Firefox(options = options)
# print("begin web driver")

# tools
def clickButton(xpath): 
    for i in range(10):
        try: 
            driver.find_element(by = "xpath", value = xpath).click()
            return  
        except: 
            sleep(0.3)
    print("unexpected press button\n\n\n")
def boxType(xpath, msg):
    for i in range(10):
        try: 
            box = driver.find_element(by = "xpath", value = xpath)
            box.send_keys(msg)
            return
        except: 
            sleep(0.3)
    print("unexpected type\n\n\n")
def getText(xpath): # and to float
    for i in range(10):
        try: 
            text = driver.find_element(by = "xpath", value = xpath).text
            try: text = float(text.replace(",", ""))
            except: pass
            return text
        except Exception as e: 
            # print(e)
            sleep(0.3)

    print("unexpected get text\n\n\n")
    return "Error"
# basic information
target = "2330"
index = 1
beginYear = 2001
endYear = 2023
delay = 0.2

# process
def read_stock_numbers_from_excel(filename):
    stock_numbers = []
    readwb = openpyxl.load_workbook(filename)
    readws = readwb.active
    for row in readws.iter_rows(values_only=True, min_row = 2):
        stock_numbers.append(row[1])
    # print("stock numbers:", *stock_numbers)
    print("got stock numbers", len(stock_numbers))
    return stock_numbers

def save_to_excel(results : list, searchYear, dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    cnt = 0
    ws.append(["公司編號"] + list(results[0].keys()))  # 添加標題行
    for i in results:
        cnt += 1
        ws.append([cnt] + list(i.values()))
    # global year
    # global season
    # print("save year", year, "season", season)
    if "results" not in os.listdir(dir): os.mkdir(dir + "/results")
    wb.save(f"{dir}/results/{searchYear}年.xlsx")

def setInformation(_beginYear, _endYear, _target):
    global target
    global index
    global beginYear
    global endYear
    target = _target
    beginYear = _beginYear
    endYear = _endYear
    results = dict()
    for year in range(beginYear, endYear + 1):
        results[year] = dict()
    # index = int(year) - beginYear + 1
    # print("set", year, season, target, index)
    return results

def login(id, password: str = "passion"):
    driver.get('https://statementdog.com/users/sign_in')
    try:
        boxType(r'//*[@id="user_email"]', id)
        boxType(r'//*[@id="user_password"]', password)
        clickButton("/html/body/div[2]/div[1]/form/div/button")
    except:
        clickButton("/html/body/div[1]/nav/div/div[2]/ul/li[3]/button/i[2]")
        clickButton("/html/body/div[1]/nav/div/div[2]/ul/li[3]/button/div/div[2]/ul/li[4]/button")
        driver.get('https://statementdog.com/users/sign_in')
        boxType(r'//*[@id="user_email"]', id)
        boxType(r'//*[@id="user_password"]', "passion")
        clickButton("/html/body/div[2]/div[1]/form/div/button")
        # driver.
    sleep(1)

def setRange(url):
    driver.get(url)
    sleep(delay)
    sleep(delay)
    clickButton(r"/html/body/div[2]/div[2]/div/div[2]/div/div[2]/div[2]/div/div[1]/div[2]/div[1]/div[1]/table/tr/td[1]/div[1]/span") # 選單
    clickButton(r"/html/body/div[2]/div[2]/div/div[2]/div/div[2]/div[2]/div/div[1]/div[2]/div[1]/div[1]/table/tr/td[1]/ul/li[4]") # 自訂
    #2001
    clickButton(r"/html/body/div[2]/div[2]/div/div[2]/div/div[2]/div[2]/div/div[1]/div[2]/div[1]/div[1]/table/tr/td[1]/div[2]/div[1]/select/option[01]")
    #2023
    clickButton(r"/html/body/div[2]/div[2]/div/div[2]/div/div[2]/div[2]/div/div[1]/div[2]/div[1]/div[1]/table/tr/td[1]/div[2]/div[2]/select/option[23]")
    clickButton(r"/html/body/div[2]/div[2]/div/div[2]/div/div[2]/div[2]/div/div[1]/div[2]/div[1]/div[1]/table/tr/td[1]/div[2]/div[3]/div")

    clickButton(r"/html/body/div[2]/div[2]/div/div[2]/div/div[2]/div[2]/div/div[1]/div[2]/div[1]/div[1]/table/tr/td[3]/ul/li[2]")

def getPrice(results: dict):
    print("Get Price:", target)
    global beginYear
    global endYear
    for year in range(beginYear, endYear + 1):
        try:
            stock = Stock(str(target))
            results[year]["股價"] = stock.fetch(year, 6)[0].close
        except:
            results[year]["股價"] = "Error"
    return results

def getCol(line: str) -> int:
    for i in range(2, 30):
        v = getText(f"/html/body/div[2]/div[2]/div/div[2]/div/div[2]/div[2]/div/div[1]/div[2]/div[2]/div[1]/ul/li[1]/table/tr[{i}]/td")
        if line == v:
            return i
        elif v == "Error": break
    return 30

def getData(results: dict, target: str, url: str, factors: list) -> list:
    print(f"Getting data for {target}, factors: {factors}")
    setRange(url)
    sleep(delay)
    global beginYear
    global endYear
    for year in range(beginYear, endYear + 1):
        index = int(year) - beginYear + 1
        for factor in factors:
            # print(year, beginYear, results)
            if year == beginYear or results[beginYear][factor] != "Error":
                results[year][factor] = getText(f"/html/body/div[2]/div[2]/div/div[2]/div/div[2]/div[2]/div/div[1]/div[2]/div[2]/div[1]/ul/li[2]/table/tr[{getCol(factor)}]/td[{index}]")
            else:
                results[year][factor] = "Error"
    return results

def getIncome(results: dict):
    return getData(results, target, f"https://statementdog.com/analysis/{target}/income-statement", ["營收", "營業利益", "稅後淨利"])

def getCash(results: dict):
    return getData(results, target, f"https://statementdog.com/analysis/{target}/cash-flow-statement", ["營業現金流", "投資現金流"])

def getRoeRoa(results: dict):
    return getData(results, target, f"https://statementdog.com/analysis/{target}/roe-roa", ["ROE", "ROA"])

def getDebt(results: dict):
    return getData(results, target, f"https://statementdog.com/analysis/{target}/liabilities-and-equity", ["總負債", "流動負債"])

def getAssets(results: dict):
    return getData(results, target, f"https://statementdog.com/analysis/{target}/assets", ["固定資產", "總資產", "流動資產"])

def getEps(results: dict):
    return getData(results, target, f"https://statementdog.com/analysis/{target}/eps", ["EPS"])



    
def safe_divide(numerator, denominator):
    numerator = str(numerator)
    denominator = str(denominator)
    try: return float(numerator.replace(",", "")) / float(denominator.replace(",", ""))
    except: return "Error"

def safe_subtract(a, b):
    a = str(a)
    b = str(b)
    try: return float(a.replace(",", "")) - float(b.replace(",", ""))
    except: return "Error"

def getOther(results):
    print("Get Other:", target)
    global beginYear
    global endYear
    
    for year in range(beginYear, endYear + 1):
        results[year]["股數"] = safe_divide(results[year]["稅後淨利"], results[year]["EPS"])
        results[year]["淨資產"] = safe_subtract(results[year]["總資產"], results[year]["總負債"])
        results[year]["每股淨值"] = safe_divide(results[year]["淨資產"], results[year]["股數"])
        results[year]["淨流動資產"] = safe_subtract(results[year]["流動資產"], results[year]["總負債"])
        results[year]["長期負債"] = safe_subtract(results[year]["總負債"], results[year]["流動負債"])
        results[year]["營業利率"] = safe_divide(results[year]["稅後淨利"], results[year]["營收"])
        results[year]["自由資本比率"] = safe_divide(results[year]["固定資產"], results[year]["總資產"])
        results[year]["負債除以本期損益"] = safe_divide(results[year]["總負債"], results[year]["營業利益"])
        results[year]["本益比"] = safe_divide(results[year]["EPS"], results[year]["股價"])
        results[year]["股價淨值比"] = safe_divide(results[year]["股價"], results[year]["每股淨值"])
    return results

            

# getInformation()
# getDebt()
# getIncome()
# getCash()
# getRoeRoa()
# getAssets()
# getPrice()


# driver.quit()    