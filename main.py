from getData import *
import openpyxl


def read_stock_numbers_from_excel(filename):
    stock_numbers = []
    readwb = openpyxl.load_workbook(filename)
    readws = readwb.active
    for row in readws.iter_rows(values_only=True):
        stock_numbers.append(row[2])
    # print("stock numbers:", *stock_numbers)
    print("got stock numbers", len(stock_numbers))
    return stock_numbers


def save_to_excel(results : list, searchYear, searchSeason):
    wb = openpyxl.Workbook()
    ws = wb.active
    cnt = 0
    ws.append(["公司編號"] + list(results[0].keys()))  # 添加標題行
    for i in results:
        cnt += 1
        ws.append([cnt] + list(i.values()))
    # global year
    # global season
    # print("save year", year, "season", season)
    wb.save(f"./results/{searchYear}年第{searchSeason}季.xlsx")

def search(beginYear, endYear):
    # l.clear()
    # global wb
    # wb = openpyxl.Workbook()
    # global ws
    # ws = wb.active
    stock_numbers = read_stock_numbers_from_excel("stocks.xlsx")
    cnt = 1
    login("h1110539@stu.wghs.tp.edu.tw")
    finalResult = {}
    for year in range(beginYear, endYear + 1):
        finalResult[year] = dict()
        for season in range(1, 5):
            finalResult[year][season] = list()
    
    for stock in stock_numbers:
        # print("目前進度:", searchYear, searchSeason, cnt)
        cnt += 1
        results = setInformation(beginYear, endYear, stock)
        results = getPrice(results)
        results = getIncome(results)
        results = getCash(results)
        results = getRoeRoa(results)
        results = getDebt(results)
        results = getAssets(results)
        results = getEps(results)
        results = getOther(results)
        print(stock, results)
        for year in range(beginYear, endYear + 1):
            for season in range(1, 5):
                finalResult[year][season].append(results[year][season])
    
    for year in range(beginYear, endYear + 1):
        for season in range(1, 5):
            save_to_excel(finalResult[year][season], year, season)
    
# fast search
# import sys

# _year = int(sys.argv[1])
# _month = int(sys.argv[2])
# print("search ", _year, _month)
# search(_year, _month)
                
search(2004, 2023)
driver.quit()
