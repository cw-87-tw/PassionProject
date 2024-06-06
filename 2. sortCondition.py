import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import os
try:
    import openpyxl
except:
    os.system("pip install openpyxl")
import tkinter.messagebox as mb

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("股票分析工具")

        self.begin_year_default = 2013
        self.end_year_default = 2022
        self.numbers_default = 5
        self.default_dir = os.getcwd()

        self.label1 = ttk.Label(root, text="開始年份:")
        self.label1.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.begin_year_entry = ttk.Entry(root)
        self.begin_year_entry.grid(row=0, column=1, padx=5, pady=5)
        self.begin_year_entry.insert(0, str(self.begin_year_default))

        self.label2 = ttk.Label(root, text="結束年份:")
        self.label2.grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.end_year_entry = ttk.Entry(root)
        self.end_year_entry.grid(row=1, column=1, padx=5, pady=5)
        self.end_year_entry.insert(0, str(self.end_year_default))

        self.label3 = ttk.Label(root, text="購買因素:")
        self.label3.grid(row=2, column=0, padx=5, pady=5, sticky="w")

        self.selected_factors = tk.StringVar(value=[])

        self.factors_listbox = tk.Listbox(root, selectmode=tk.MULTIPLE, listvariable=self.selected_factors, height=5)
        self.factors_listbox.grid(row=2, column=1, padx=5, pady=5)

        self.add_factor_button = ttk.Button(root, text="新增/刪除因素", command=self.open_factor_window)
        self.add_factor_button.grid(row=2, column=2, padx=5, pady=5)

        self.label4 = ttk.Label(root, text="每年購買股票數量:")
        self.label4.grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.numbers_entry = ttk.Entry(root)
        self.numbers_entry.grid(row=3, column=1, padx=5, pady=5)
        self.numbers_entry.insert(0, str(self.numbers_default))

        self.label5 = ttk.Label(root, text="資料目錄:")
        self.label5.grid(row=4, column=0, padx=5, pady=5, sticky="w")
        self.save_path_entry = ttk.Entry(root, state="readonly")
        self.save_path_entry.grid(row=4, column=1, padx=5, pady=5)
        self.save_path_entry.insert(0, self.default_dir)
        self.browse_button = ttk.Button(root, text="瀏覽...", command=self.browse)
        self.browse_button.grid(row=4, column=2, padx=5, pady=5)

        self.run_button = ttk.Button(root, text="執行", command=self.run)
        self.run_button.grid(row=5, column=1, padx=5, pady=5)

        self.factors = []

    def browse(self):
        path = filedialog.askdirectory()
        self.save_path_entry.configure(state="normal")
        self.save_path_entry.delete(0, tk.END)
        self.save_path_entry.insert(0, path)
        self.save_path_entry.configure(state="readonly")

    def open_factor_window(self):
        factor_window = tk.Toplevel(self.root)
        factor_window.title("新增/刪除因素")

        label = ttk.Label(factor_window, text="因素名稱:")
        label.grid(row=0, column=0, padx=5, pady=5)

        self.new_factor_entry = ttk.Entry(factor_window)
        self.new_factor_entry.grid(row=0, column=1, padx=5, pady=5)

        label = ttk.Label(factor_window, text="倍數:")
        label.grid(row=1, column=0, padx=5, pady=5)

        self.new_factor_multiple_entry = ttk.Entry(factor_window)
        self.new_factor_multiple_entry.grid(row=1, column=1, padx=5, pady=5)

        add_button = ttk.Button(factor_window, text="新增", command=self.add_factor)
        add_button.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

        remove_button = ttk.Button(factor_window, text="刪除所選因素", command=self.remove_factor)
        remove_button.grid(row=3, column=0, columnspan=2, padx=5, pady=5)

        self.current_factors_listbox = tk.Listbox(factor_window, selectmode=tk.SINGLE, height=5)
        self.current_factors_listbox.grid(row=4, column=0, columnspan=2, padx=5, pady=5)

        for factor in self.factors:
            self.current_factors_listbox.insert(tk.END, factor)

    def add_factor(self):
        factor_name = self.new_factor_entry.get()
        factor_multiple = self.new_factor_multiple_entry.get()
        factor = (factor_name, factor_multiple)
        if factor not in self.factors:
            self.factors.append(factor)
            self.current_factors_listbox.insert(tk.END, factor)
            self.factors_listbox.insert(tk.END, factor)
        self.new_factor_entry.delete(0, tk.END)
        self.new_factor_multiple_entry.delete(0, tk.END)

    def remove_factor(self):
        selection = self.current_factors_listbox.curselection()
        if selection:
            index = selection[0]
            factor = self.current_factors_listbox.get(index)
            self.factors.remove(factor)
            self.current_factors_listbox.delete(index)
            self.factors_listbox.delete(index)

    def run(self):
        begin_year = int(self.begin_year_entry.get())
        end_year = int(self.end_year_entry.get())
        numbers = int(self.numbers_entry.get())
        # print(self.factors_listbox.get(0))
        factors = [(self.factors_listbox.get(idx)[0], float(self.factors_listbox.get(idx)[1])) for idx in range(self.factors_listbox.size())]
        dir = self.save_path_entry.get()
        results = getData(begin_year, end_year, factors, dir, numbers)
        run(results, begin_year, end_year, factors, dir, numbers)

def getData(beginYear, endYear, factors, dir, numbers) -> dict:
    """
    the function can get data from the results crawled on the Net
    the workbook must be located at the desinated directory(./results/)
    it would get all the data in the following format
    results = { 
        beginYear: {
            1: {
                factor1: 123,
                factor2: "Error"...
            },
            2: {
            
            }...
        }...
    }
    """
    results = dict()
    for year in range(beginYear, endYear + 1):
        results[year] = dict()
        readwb = openpyxl.load_workbook(f"{dir}/results/{year}年.xlsx")
        readws = readwb.active

        for col in readws.iter_cols(values_only=True, min_col=2):
            factor = col[0]
            col = col[1:]
            for idx, i in enumerate(col, 1):
                if results[year].get(idx) == None: results[year][idx] = dict()
                try: results[year][idx][factor] = float(i)
                except: results[year][idx][factor] = "Error"
    return results

def customSort(results: dict, year: int, factors: list, numbers: int) -> list:
    """
    This function can return the best stock numbers to buy
    parameters:
        results: results from getData function
        year: the target year
        factors: a list of factor, each factor is formatted like (factorName, multiple)
    the sorting function will calculate the points for each stock, and return the top {numbers} stocks
    the points would be calculated by finding the index, the from {len(stocks)} to 1, multiply the {multiple} of the factor
    """
    dataList = [(stock, data) for stock, data in results[year].items()]
    
    indexed = dict()
    for factor, multiple in factors:
        tmp = list()
        for stock, data in dataList:
            if type(data[factor]) != str: tmp.append(data[factor])
        indexed[factor] = sorted(tmp)

    def customKey(item: tuple):
        key = 0
        for factor, multiple in factors:
            # print(key, indexed[factor].index(item[1][factor]) * multiple)
            if type(item[1][factor]) != str and type(item[1]["股價"]) != str: key += indexed[factor].index(item[1][factor]) * multiple
            else: key += (-1e20) * abs(multiple)
        return key

    dataList.sort(key=customKey, reverse=True)
    buyList = list()
    for stock, data in dataList[:numbers]:
        buyList.append(stock)
    return buyList

def run(results, beginYear, endYear, factors, dir, numbers):
    try:
        output = Sheet(beginYear)
        for year in range(beginYear, endYear + 1):
            buyList = customSort(results, year, factors, numbers)
            output.addData(buyList, year)

        outputList = [f"{factor} x {multiple}" for factor, multiple in factors]

        saveDir = "buyTargets"
        if saveDir not in os.listdir(dir): os.mkdir(dir + "/" + saveDir)
    
        output.save(f"{dir}/{saveDir}/購買標的({','.join(outputList)})")
        print("檔案成功儲存:", f"{dir}/{saveDir}/購買標的({','.join(outputList)}).xlsx")
        mb.showinfo("成功", f"檔案儲存成功: {dir}/{saveDir}/購買標的({','.join(outputList)}).xlsx")
    except Exception as e:
        print("失敗，錯誤訊息如下:")
        print(e)
        mb.showerror("失敗", f"檔案儲存失敗，錯誤訊息如下:\n{e}")


class Sheet:
    def __init__(self, beginYear):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.beginYear = beginYear

    def addData(self, buyList: list, year: int) -> None:
        if year == self.beginYear: self.ws.append(["年份", "以下為購買的對象編號"])
        self.ws.append([year] + buyList)
    
    def save(self, filename: str):
        self.wb.save(f"{filename}.xlsx")

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
