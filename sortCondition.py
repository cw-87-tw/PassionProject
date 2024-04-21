import openpyxl

beginYear = 2014 
beginYear = int(input("beginYear:"))
endYear = 2022
endYear = int(input("endYear:"))

def getData() -> dict:
    results = dict()
    for year in range(beginYear, endYear + 1):
        results[year] = dict()
        readwb = openpyxl.load_workbook(f"./results/{year}年.xlsx")
        readws = readwb.active

        for col in readws.iter_cols(values_only = True, min_col = 2):
            factor = col[0]
            col = col[1:]
            for idx, i in enumerate(col, 1): # 到時候可能不能用 enumerate 因為要用真的股票編號
                if results[year].get(idx) == None: results[year][idx] = dict()
                results[year][idx][factor] = i if type(i) != str else 1e20
                # print(idx, prices[idx])
    return results

def customSort(results: dict, year: int, numbers: int = 5) -> list:
    dataList = [(stock, data) for stock, data in results[year].items()]
    def customKey(item: tuple):
        return item[1]["股價"]
    dataList.sort(key = customKey, reverse = False)
    buyList = list()
    for stock, data in dataList[:numbers]:
        buyList.append(stock)
    return buyList

wb = openpyxl.Workbook()
ws = wb.active
def addData(buyList: list, year: int) -> None:
    if year == beginYear: ws.append(["年份", "以下為購買的對象編號"])
    ws.append([year] + buyList)


results = getData()
print(results)
for year in range(beginYear, endYear + 1):
    buyList = customSort(results, year)
    addData(buyList, year)


wb.save("購買標的.xlsx")