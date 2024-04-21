import openpyxl

beginYear = 2013
# beginYear = int(input("beginYear:"))
endYear = 2022
# endYear = int(input("endYear:"))
factor = "營收"
highest = True
numbers = 5

def getData() -> dict:
    results = dict()
    for year in range(beginYear, endYear + 1):
        results[year] = dict()
        readwb = openpyxl.load_workbook(f"./results/{year}年.xlsx")
        readws = readwb.active

        for col in readws.iter_cols(values_only = True, min_col = 2):
            factor = col[0]
            col = col[1:]
            for idx, i in enumerate(col, 1): # 到時候可能不能用 enumerate 因為要用真的股票編號
                if results[year].get(idx) == None: results[year][idx] = dict()
                try: results[year][idx][factor] = float(i)
                except: results[year][idx][factor] = "Error"
                # print(idx, prices[idx])
    return results

def customSort(results: dict, year: int, factor: str, highest: bool = False, numbers: int = 5,) -> list: # number: 要挑出最好的幾個?
    dataList = [(stock, data) for stock, data in results[year].items()]
    def customKey(item: tuple): # 用來排序的判準
        if type(item[1][factor]) != float or item[1]["股價"] == "Error":
            # print(item)
            if highest: return -1e20
            else: return 1e20
        return item[1][factor]
    dataList.sort(key = customKey, reverse = highest) # reverse: False 小到大, True 大到小
    buyList = list()
    for stock, data in dataList[:numbers]:
        buyList.append(stock)
    return buyList

wb = openpyxl.Workbook()
ws = wb.active
def addData(buyList: list, year: int) -> None:
    if year == beginYear: ws.append(["年份", "以下為購買的對象編號"])
    ws.append([year] + buyList)


results = getData()
# print(results)
for year in range(beginYear, endYear + 1):
    buyList = customSort(results, year, factor = factor, highest = highest, numbers = numbers)
    addData(buyList, year)


wb.save(f"./buyResults/購買標的({factor}).xlsx")




buySheet = f"./buyResults/購買標的({factor}).xlsx"
money = 5000000

wb = openpyxl.load_workbook(buySheet)
ws = wb.active

allTargets = dict()
for row in ws.iter_rows(values_only=True, min_row = 2):
    allTargets[row[0]] = list()
    for cell in row[1:]:
        if cell != None: allTargets[row[0]].append(cell)

print(allTargets)

def getPrices(year: int) -> dict:
    readwb = openpyxl.load_workbook(f"./results/{year}年.xlsx")
    readws = readwb.active
    idx = 1
    prices = dict()
    for col in readws.iter_cols(values_only = True, min_col = 2, max_col = 2):
        for i in col[1:]:
            prices[idx] = i if type(i) != str else 1e20
            # print(idx, prices[idx])
            idx += 1
    return prices

hold = [] # (number, 股數)
wb = openpyxl.Workbook()
ws = wb.active

for year, targets in allTargets.items():
    prices = getPrices(year) # get this year's prices

    # sell old
    for no, shares in hold:
        if prices[no] == 1e20: raise Exception(f"{year}年無法賣出此股票(可能因為下市等因素)\n錯誤編號為{no}")
        money += shares * prices[no]
    hold.clear()

    ws.append([f"{year}年"])
    ws.append(["原有資金", money])
    ws.append([])
    ws.append(["購買編號", "股價", "股數", "總價"])

    # buy new
    perMoney = money / len(targets)
    for stock in targets:
        shares = perMoney // prices[stock]
        money -= prices[stock] * shares
        ws.append([stock, prices[stock], shares, prices[stock] * shares])
        print(f"{year}購買了{stock}，買了{shares}股")
        hold.append((stock, shares))

    ws.append([])
    ws.append(["剩餘資金", money])
    ws.append([])
    ws.append([])
    ws.append([])


lastYear = list(allTargets.keys())[-1] + 1
prices = getPrices(lastYear) # get this year's prices
# sell old
for no, shares in hold:
    if prices[no] == 1e20: raise Exception(f"{lastYear}年無法賣出此股票(可能因為下市等因素)\n錯誤編號為{no}")
    money += shares * prices[no]
hold.clear()

ws.append(["最終資金", money])
    
wb.save(f"./buyResults/購買分析結果({factor}).xlsx")